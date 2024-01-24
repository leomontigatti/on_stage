import io

from django.contrib import admin, messages
from django.contrib.admin import SimpleListFilter
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.translation import gettext
from django.utils.translation import gettext_lazy as _
from django.utils.translation import ngettext
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from weasyprint import HTML

from choreography.forms import (
    AwardAdminForm,
    AwardFormSet,
    AwardInlineForm,
    DiscountAdminForm,
    DiscountFormSet,
    PaymentAdminForm,
    PaymentFormSet,
    ScoreInlineForm,
)
from choreography.models import Award, Choreography, Discount, Feedback, Payment, Score

OSUser = get_user_model()


class OrderNumberFilter(SimpleListFilter):
    title = _("Order number")
    parameter_name = "order_number"

    def lookups(self, request, model_admin):
        return (
            ("ordered", _("Ordered")),
            ("not_ordered", _("Not ordered")),
        )

    def queryset(self, request, queryset):
        if self.value() == "ordered":
            return queryset.exclude(order_number=None)
        elif self.value() == "not_ordered":
            return queryset.filter(order_number=None)
        else:
            return queryset


class FullyPaidFilter(SimpleListFilter):
    title = _("Fully paid")
    parameter_name = "fully_paid"

    def lookups(self, request, model_admin):
        return (
            ("true", _("Yes")),
            ("false", _("No")),
        )

    def queryset(self, request, queryset):
        paid_choreographies = [
            choreography.pk for choreography in queryset if choreography.fully_paid
        ]
        if self.value() == "true":
            return queryset.filter(pk__in=paid_choreographies)
        elif self.value() == "false":
            return queryset.exclude(pk__in=paid_choreographies)
        else:
            return queryset


def assign_judge(judge):
    name = f"assign_judge_{judge}"

    def assign_judge_action(modeladmin, request, queryset):
        filtered_qs = queryset.filter(event__end_date__gte=timezone.now().date())
        count = filtered_qs.count()
        for choreography in filtered_qs:
            Score.objects.get_or_create(choreography=choreography, judge=judge)
        message = ngettext(
            "%(count)d choreography was assigned to %(judge)s",
            "%(count)d choreographies were assigned to %(judge)s",
            count,
        ) % {"count": count, "judge": judge}
        modeladmin.message_user(request, message, messages.SUCCESS)

    return (
        name,
        (
            assign_judge_action,
            name,
            _("Assign to judge") + f": {judge}",
        ),
    )


@admin.register(Choreography)
class ChoreographyAdmin(admin.ModelAdmin):
    list_display = ["order_number", "id", "name", "academy", "category", "dance_mode"]
    list_filter = [
        "event",
        OrderNumberFilter,
        FullyPaidFilter,
        "dance_mode",
        "category__type",
        "category__name",
    ]
    list_display_links = ["id", "name"]
    search_fields = ["id", "event__name", "name", "academy__name"]
    search_help_text = _("Search by PK, event, name or academy.")
    show_facets = admin.ShowFacets.ALWAYS

    actions = [
        "show_awards",
        "hide_awards",
        "manage_payments",
        "set_order_number",
        "export_pdf",
        "export_event_excel",
        "export_accounting_excel",
    ]

    @admin.action(description=_("Show awards"))
    def show_awards(self, request, queryset):
        queryset.update(show_awards=True)
        self.message_user(
            request,
            _("Awards for the selected choreographies are now shown."),
            messages.SUCCESS,
        )

    @admin.action(description=_("Hide awards"))
    def hide_awards(self, request, queryset):
        queryset.update(show_awards=False)
        self.message_user(
            request,
            _("Awards for the selected choreographies are now hidden."),
            messages.SUCCESS,
        )

    @admin.action(description=_("Manage payments"))
    def manage_payments(self, request, queryset):
        context = {
            "choreographies_list": queryset.filter(
                event__end_date__gte=timezone.now().date()
            ).order_by("-pk"),
            "has_permission": request.user.groups.filter(name="Admin").exists(),
            "site_url": "/",
            "title": _("Manage payments"),
        }
        return render(
            request, "choreography/choreography_payment.html", context=context
        )

    @admin.action(description=_("Set order number"))
    def set_order_number(self, request, queryset):
        context = {
            "choreographies_list": queryset.filter(
                event__end_date__gte=timezone.now().date()
            ).order_by("-pk"),
            "has_permission": request.user.groups.filter(name="Admin").exists(),
            "site_url": "/",
            "title": _("Set order number"),
        }
        return render(
            request, "choreography/choreography_set_order_number.html", context=context
        )

    @admin.action(description=_("Export to PDF"))
    def export_pdf(self, request, queryset):
        context = {
            "model": Choreography,
            "choreographies_list": queryset,
        }
        html_string = render_to_string(
            "choreography/choreography_export_pdf.html", context, request
        )
        pdf_file = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf_file, content_type="application/pdf")
        return response

    @admin.action(description=_("Export to Excel | Event"))
    def export_event_excel(self, request, queryset):
        excel_file = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = gettext("Choreographies list")

        headers = [
            gettext("Order number"),
            gettext("Category"),
            gettext("Dance mode"),
            gettext("Academy"),
            gettext("State"),
            gettext("Choreography name"),
            gettext("Professors"),
            gettext("Dancers"),
            gettext("Dancers amount"),
        ]

        title_cell = worksheet.cell(row=1, column=1)
        count = queryset.count()
        message = ngettext(
            "%(count)d selected choreography",
            "%(count)d selected choreographies",
            count,
        ) % {
            "count": count,
        }
        title_cell.value = message
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.merge_cells(
            start_column=1, end_column=len(headers), start_row=1, end_row=1
        )

        headers_row = 2
        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=headers_row, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.font = Font(bold=True)

        current_row = headers_row + 1
        for choreography in queryset:
            professors = [
                professor.__str__() for professor in choreography.professors.all()
            ]
            dancers = [dancer.__str__() for dancer in choreography.dancers.all()]

            row = [
                choreography.order_number or "-",
                choreography.category.__str__(),
                choreography.dance_mode.__str__(),
                choreography.academy.__str__(),
                choreography.academy.state,
                choreography.name,
                "\n".join(professors),
                "\n".join(dancers),
                choreography.dancers.count(),
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(vertical="center")

            worksheet.row_dimensions[current_row].height = (
                max(len(professors), len(dancers)) * 15
            )

            current_row += 1

        for column in worksheet.columns:
            for col in column:
                if hasattr(col, "column_letter"):
                    worksheet.column_dimensions[col.column_letter].width = 20

        workbook.save(excel_file)
        response = HttpResponse(
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response[
            "Content-Disposition"
        ] = "attachment; filename=Choreographies list.xlsx"
        return response

    @admin.action(description=_("Export to Excel | Accounting"))
    def export_accounting_excel(self, request, queryset):
        excel_file = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = gettext("Choreographies list")

        headers = [
            gettext("Choreography ID"),
            gettext("Category"),
            gettext("Dance mode"),
            gettext("Academy"),
            gettext("Choreography name"),
            gettext("Dancers amount"),
            gettext("Price per dancer"),
            gettext("Discounts"),
            gettext("Total"),
            gettext("Paid amount"),
            gettext("Balance"),
        ]

        ordered_queryset = queryset.order_by("pk")

        total_price_amount, total_paid_amount = 0, 0
        for choreography in ordered_queryset:
            total_price_amount += choreography.total_price
            if choreography.payments.exists():
                for payment in choreography.payments.all():
                    total_paid_amount += payment.amount

        title_cell = worksheet.cell(row=1, column=1)
        count = queryset.count()
        message = ngettext(
            "%(count)d selected choreography - Total $ %(price)s - Paid amount $ %(paid)s",
            "%(count)d selected choreographies - Total $ %(price)s - Paid amount $ %(paid)s",
            count,
        ) % {
            "count": count,
            "price": total_price_amount,
            "paid": total_paid_amount,
        }
        title_cell.value = message
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.merge_cells(
            start_column=1, end_column=len(headers), start_row=1, end_row=1
        )

        headers_row = 2
        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=headers_row, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.font = Font(bold=True)

        current_row = headers_row + 1
        for choreography in ordered_queryset:
            total_discounts = 0
            if choreography.discounts.exists():
                for discount in choreography.discounts.all():
                    total_discounts -= discount.amount

            row = [
                choreography.pk,
                choreography.category.__str__(),
                choreography.dance_mode.__str__(),
                choreography.academy.__str__(),
                choreography.name,
                choreography.dancers.count(),
                choreography.price.amount,
                total_discounts,
                choreography.total_price,
                -choreography.paid_amount,
                choreography.balance,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(vertical="center")

            current_row += 1

        for column in worksheet.columns:
            for col in column:
                if hasattr(col, "column_letter"):
                    worksheet.column_dimensions[col.column_letter].width = 20

        workbook.save(excel_file)
        response = HttpResponse(
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response[
            "Content-Disposition"
        ] = "attachment; filename=Choreographies list.xlsx"
        return response

    def get_actions(self, request):
        actions = super().get_actions(request)
        judges_list = OSUser.objects.filter(groups__name="Judge").order_by("first_name")
        return {**actions, **dict(assign_judge(judge) for judge in judges_list)}

    fieldsets = (
        (
            _("General"),
            {
                "classes": ("collapse",),
                "fields": (
                    "academy",
                    ("event", "schedule"),
                    ("name", "duration"),
                    ("dance_mode", "category"),
                    "professors",
                    "dancers",
                    "music_track",
                ),
            },
        ),
        (
            None,
            {
                "fields": (
                    ("order_number", "average_score"),
                    ("is_locked", "is_disqualified", "show_awards"),
                    ("price", "total_price", "balance"),
                    ("deposit_paid", "fully_paid"),
                )
            },
        ),
    )

    filter_horizontal = ["professors", "dancers"]

    def get_readonly_fields(self, request, obj):
        readonly_fields = [
            "total_price",
            "average_score",
            "deposit_paid",
            "fully_paid",
            "balance",
        ]
        if not request.user.is_superuser and obj:
            readonly_fields.extend([field.name for field in self.model._meta.fields])
            readonly_fields.extend(["dancers", "professors"])
            if not obj.event.ended:
                readonly_fields.remove("order_number")
                readonly_fields.remove("is_locked")
            readonly_fields.remove("is_disqualified")
            readonly_fields.remove("show_awards")
        return readonly_fields

    class PaymentInline(admin.TabularInline):
        model = Payment
        fk_name = "choreography"
        extra = 0
        formset = PaymentFormSet

        def has_change_permission(self, request, obj=None):
            return request.user.is_superuser or obj and not obj.event.ended

        def has_add_permission(self, request, obj=None):
            return (
                request.user.is_superuser
                or obj
                and not obj.event.ended
                and obj.balance > 0
            )

        def has_delete_permission(self, request, obj=None):
            return request.user.is_superuser or obj and not obj.event.ended

    class DiscountInline(admin.TabularInline):
        model = Discount
        fk_name = "choreography"
        extra = 0
        formset = DiscountFormSet

        def has_change_permission(self, request, obj=None):
            return request.user.is_superuser or obj and not obj.event.ended

        def has_add_permission(self, request, obj=None):
            return (
                request.user.is_superuser
                or obj
                and not obj.event.ended
                and obj.balance > 0
            )

        def has_delete_permission(self, request, obj=None):
            return request.user.is_superuser or obj and not obj.event.ended

    class AwardInline(admin.TabularInline):
        model = Award
        fk_name = "choreography"
        extra = 0
        form = AwardInlineForm
        formset = AwardFormSet

        def has_delete_permission(self, request, obj=None):
            return request.user.is_superuser

        def has_add_permission(self, request, obj=None):
            return request.user.is_superuser or not obj.is_disqualified

    class ScoreInline(admin.TabularInline):
        model = Score
        fk_name = "choreography"
        extra = 0
        form = ScoreInlineForm

        def has_change_permission(self, request, obj=None):
            return request.user.is_superuser or obj and not obj.event.ended

    inlines = [PaymentInline, DiscountInline, AwardInline, ScoreInline]

    @admin.display(description=_("Total price"))
    def total_price(self, obj):
        return f"$ {obj.total_price}"

    @admin.display(description=_("Balance"))
    def balance(self, obj):
        return f"$ {obj.balance}"

    @admin.display(boolean=True, description=_("Fully paid"))
    def fully_paid(self, obj):
        return obj.fully_paid

    @admin.display(boolean=True, description=_("Deposit paid"))
    def deposit_paid(self, obj):
        return obj.deposit_paid


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    list_display = ["date", "choreography_name", "academy", "amount", "payment_method"]
    list_filter = ["choreography__event", "payment_method", "date"]
    list_display_links = ["choreography_name"]
    search_fields = [
        "choreography__id",
        "choreography__academy__name",
        "date",
        "choreography__name",
    ]
    search_help_text = _("Search by academy, date, choreography ID or name.")
    show_facets = admin.ShowFacets.ALWAYS

    actions = ["export_excel"]

    @admin.display(description=_("Choreography name"))
    def choreography_name(self, obj):
        return obj.choreography.name

    @admin.display(description=_("Academy"))
    def academy(self, obj):
        return obj.choreography.academy

    @admin.action(description=_("Export to Excel"))
    def export_excel(self, request, queryset):
        excel_file = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = gettext("Payments list")

        headers = [
            gettext("Date"),
            gettext("Choreography ID"),
            gettext("Choreography name"),
            gettext("Academy"),
            gettext("Amount"),
            gettext("Payment method"),
        ]

        total_amount = 0
        for payment in queryset:
            total_amount += payment.amount

        title_cell = worksheet.cell(row=1, column=1)
        count = queryset.count()
        message = ngettext(
            "%(count)d selected payment - Total $ %(amount)s",
            "%(count)d selected payments - Total $ %(amount)s",
            count,
        ) % {
            "count": count,
            "amount": total_amount,
        }
        title_cell.value = message
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.merge_cells(
            start_column=1, end_column=len(headers), start_row=1, end_row=1
        )

        headers_row = 2
        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=headers_row, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.font = Font(bold=True)

        current_row = headers_row + 1
        for payment in queryset:
            row = [
                payment.date,
                payment.choreography.pk,
                payment.choreography.name,
                payment.choreography.academy.__str__(),
                payment.amount,
                payment.get_payment_method_display(),
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(vertical="center")

            current_row += 1

        for column in worksheet.columns:
            for col in column:
                if hasattr(col, "column_letter"):
                    worksheet.column_dimensions[col.column_letter].width = 20

        workbook.save(excel_file)
        response = HttpResponse(
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=Payments list.xlsx"
        return response

    fieldsets = (
        (
            None,
            {"fields": ("choreography", "amount", "payment_method", "date")},
        ),
    )

    form = PaymentAdminForm

    def get_readonly_fields(self, request, obj):
        return ["choreography"] if obj else []

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser or obj and not obj.choreography.event.ended

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser or obj and not obj.choreography.event.ended


@admin.register(Award)
class AwardAdmin(admin.ModelAdmin):
    list_display = (
        "order_number",
        "choreography_pk",
        "choreography_name",
        "academy_name",
        "award_type",
        "average_score",
    )
    list_filter = [
        "choreography__event",
        "choreography__is_disqualified",
        "award_type",
        "choreography__dance_mode__name",
        "choreography__category__name",
    ]
    list_display_links = ["choreography_pk", "choreography_name"]
    search_fields = [
        "choreography__id",
        "choreography__academy__name",
        "choreography__name",
    ]
    search_help_text = _("Search by academy name, choreography ID or name.")
    show_facets = admin.ShowFacets.ALWAYS

    @admin.display(description=_("Choreography ID"))
    def choreography_pk(self, obj):
        return obj.choreography.pk

    @admin.display(description=_("Choreography name"))
    def choreography_name(self, obj):
        return obj.choreography.name

    @admin.display(description=_("Academy"))
    def academy_name(self, obj):
        return obj.choreography.academy

    @admin.display(description=_("Average score"))
    def average_score(self, obj):
        return obj.choreography.average_score

    @admin.display(description=_("Order number"))
    def order_number(self, obj):
        return obj.choreography.order_number

    actions = ["export_excel"]

    @admin.action(description=_("Export to Excel"))
    def export_excel(self, request, queryset):
        excel_file = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = gettext("Awards list")

        headers = [
            gettext("Order number"),
            gettext("Category"),
            gettext("Dance mode"),
            gettext("Academy"),
            gettext("State"),
            gettext("Choreography name"),
            gettext("Professors"),
            gettext("Dancers"),
            gettext("Dancers amount"),
            gettext("Average score"),
            gettext("Award"),
        ]

        title_cell = worksheet.cell(row=1, column=1)
        count = queryset.count()
        message = ngettext(
            "%(count)d selected award",
            "%(count)d selected awards",
            count,
        ) % {
            "count": count,
        }
        title_cell.value = message
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.merge_cells(
            start_column=1, end_column=len(headers), start_row=1, end_row=1
        )

        headers_row = 2
        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=headers_row, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.font = Font(bold=True)

        current_row = headers_row + 1
        for award in queryset.exclude(choreography__is_disqualified=True):
            professors = [
                professor.__str__() for professor in award.choreography.professors.all()
            ]
            dancers = [dancer.__str__() for dancer in award.choreography.dancers.all()]

            row = [
                award.choreography.order_number or "-",
                award.choreography.category.__str__(),
                award.choreography.dance_mode.__str__(),
                award.choreography.academy.name,
                award.choreography.academy.state,
                award.choreography.name,
                "\n".join(professors),
                "\n".join(dancers),
                len(dancers),
                award.choreography.average_score,
                award.award_type.__str__(),
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(vertical="center")

            worksheet.row_dimensions[current_row].height = (
                max(len(professors), len(dancers)) * 15
            )

            current_row += 1

        for column in worksheet.columns:
            for col in column:
                if hasattr(col, "column_letter"):
                    worksheet.column_dimensions[col.column_letter].width = 20

        workbook.save(excel_file)
        response = HttpResponse(
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=Awards list.xlsx"
        return response

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "choreography",
                    "assigned_by",
                    "award_type",
                )
            },
        ),
    )

    form = AwardAdminForm

    def get_readonly_fields(self, request, obj):
        return ["choreography"] if obj else []

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser or obj and not obj.assigned_by_id == 1

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser or obj and not obj.assigned_by_id == 1


@admin.register(Score)
class ScoreAdmin(admin.ModelAdmin):
    list_display = [
        "choreography_order_number",
        "choreography_pk",
        "choreography_name",
        "academy_name",
        "judge",
        "value",
        "is_locked",
    ]
    list_filter = [
        "choreography__event",
        "judge",
        "choreography__category__name",
        "choreography__dance_mode__name",
    ]
    list_display_links = ["choreography_pk", "choreography_name"]
    search_fields = [
        "choreography__academy__name",
        "choreography__id",
        "choreography__name",
    ]
    search_help_text = _("Search by PK, academy or choreography name.")
    actions = ["set_is_locked", "set_is_unlocked"]
    show_facets = admin.ShowFacets.ALWAYS

    @admin.display(description=_("Order number"))
    def choreography_order_number(self, obj):
        return obj.choreography.order_number

    @admin.display(description="Choreography ID")
    def choreography_pk(self, obj):
        return obj.choreography.pk

    @admin.display(description=_("Choreography name"))
    def choreography_name(self, obj):
        return obj.choreography.name

    @admin.display(description=_("Academy"))
    def academy_name(self, obj):
        return obj.choreography.academy

    @admin.action(description=_("Lock selected scores."))
    def set_is_locked(self, request, queryset):
        filtered_qs = queryset.filter(event__end_date__gte=timezone.now().date())
        locked = filtered_qs.update(is_locked=True)
        message = ngettext(
            "%(locked)d score was locked.",
            "%(locked)d scores were locked.",
            locked,
        ) % {
            "locked": locked,
        }
        self.message_user(request, message, messages.SUCCESS)

    @admin.action(description=_("Unlock selected scores."))
    def set_is_unlocked(self, request, queryset):
        filtered_qs = queryset.filter(event__end_date__gte=timezone.now().date())
        unlocked = filtered_qs.update(is_locked=False)
        message = ngettext(
            "%(unlocked)d score was unlocked.",
            "%(unlocked)d scores were unlocked.",
            unlocked,
        ) % {
            "unlocked": unlocked,
        }
        self.message_user(request, message, messages.SUCCESS)

    @admin.display(description=_("Feedback"))
    def feedback(self, obj):
        return obj.feedback

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "choreography",
                    ("judge", "feedback"),
                    ("value", "is_locked"),
                )
            },
        ),
    )

    def get_readonly_fields(self, request, obj):
        if obj:
            if request.user.is_superuser:
                return ["choreography", "feedback"]
            readonly_fields = ["choreography", "judge", "value", "amount", "feedback"]
            if obj.choreography.event.ended:
                readonly_fields.append("is_locked")
            return readonly_fields
        return []


@admin.register(Discount)
class DiscountAdmin(admin.ModelAdmin):
    list_display = ["choreography_pk", "choreography_name", "academy_name", "amount"]
    list_filter = [
        "choreography__event",
        "choreography__academy",
    ]
    list_display_links = ["choreography_name", "choreography_pk"]
    search_fields = [
        "choreography__id",
        "choreography__academy__name",
        "choreography__name",
    ]
    search_help_text = _("Search by academy, choreography ID or name.")
    show_facets = admin.ShowFacets.ALWAYS

    @admin.display(description=_("Choreography ID"))
    def choreography_pk(self, obj):
        return obj.choreography.pk

    @admin.display(description=_("Choreography name"))
    def choreography_name(self, obj):
        return obj.choreography.name

    @admin.display(description=_("Academy"))
    def academy_name(self, obj):
        return obj.choreography.academy

    fieldsets = (
        (
            None,
            {"fields": ("choreography", "amount")},
        ),
    )

    form = DiscountAdminForm

    def get_readonly_fields(self, request, obj):
        return ["choreography"] if obj else []

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser or obj and not obj.choreography.event.ended

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser or obj and not obj.choreography.event.ended


@admin.register(Feedback)
class FeedbackAdmin(admin.ModelAdmin):
    list_display = ["choreography_id", "choreography_name", "academy", "judge"]
    list_filter = ["score__choreography__event", "score__judge"]
    list_display_links = ["choreography_id", "choreography_name"]
    search_fields = [
        "score__choreography__academy",
        "score__choreography_pk",
        "score__choreography_name",
    ]
    search_help_text = _("Search by academy, choreography ID or name.")
    show_facets = admin.ShowFacets.ALWAYS

    @admin.display(description=_("Choreography ID"))
    def choreography_id(self, obj):
        return obj.score.choreography.pk

    @admin.display(description=_("Choreography name"))
    def choreography_name(self, obj):
        return obj.score.choreography.name

    @admin.display(description=_("Academy"))
    def academy(self, obj):
        return obj.score.choreography.academy

    @admin.display(description=_("Judge"))
    def judge(self, obj):
        return obj.score.judge

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "score",
                    "audio_file",
                )
            },
        ),
    )

    def get_readonly_fields(self, request, obj):
        if obj:
            return ["score"]
        return []
