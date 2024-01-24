import io

from django.contrib import admin
from django.contrib.admin import SimpleListFilter
from django.http import HttpResponse
from django.utils.translation import gettext
from django.utils.translation import gettext_lazy as _
from django.utils.translation import ngettext
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from academy.models import Dancer
from seminar.forms import (
    SeminarAdminForm,
    SeminarPaymentAdminForm,
    SeminarPaymentFormSet,
)
from seminar.models import Seminar, SeminarPayment, SeminarPrice, SeminarRegistration


class IsFullFilter(SimpleListFilter):
    title = _("Is full")
    parameter_name = "is_full"

    def lookups(self, request, model_admin):
        return (
            ("true", _("Yes")),
            ("false", _("No")),
        )

    def queryset(self, request, queryset):
        full_seminars = [seminar.pk for seminar in queryset if seminar.is_full]
        if self.value() == "true":
            return queryset.filter(pk__in=full_seminars)
        elif self.value() == "false":
            return queryset.exclude(pk__in=full_seminars)
        else:
            return queryset


@admin.register(SeminarPrice)
class SeminarPriceAdmin(admin.ModelAdmin):
    list_display = [
        "type",
        "one_registration_price",
        "two_registrations_price",
        "more_registrations_price",
    ]
    list_filter = ["type"]
    list_display_links = ["type"]
    show_facets = admin.ShowFacets.ALWAYS

    fieldsets = (
        (
            None,
            {
                "fields": (
                    ("type",),
                    (
                        "one_registration_price",
                        "two_registrations_price",
                        "more_registrations_price",
                    ),
                )
            },
        ),
    )


@admin.register(Seminar)
class SeminarAdmin(admin.ModelAdmin):
    list_display = ["teacher", "special_price", "date", "time", "is_full"]
    list_filter = ["event", IsFullFilter, "teacher"]
    list_display_links = ["teacher"]
    show_facets = admin.ShowFacets.ALWAYS

    fieldsets = (
        (
            None,
            {
                "fields": (
                    ("event",),
                    ("teacher", "price"),
                    ("date", "time", "registration_end_date"),
                    ("quota", "available_space", "is_full"),
                    ("teacher_picture",),
                )
            },
        ),
    )

    readonly_fields = ["available_space", "is_full"]

    form = SeminarAdminForm

    @admin.display(description=_("Available space"))
    def available_space(self, obj):
        return f"{obj.available_space}"

    @admin.display(boolean=True, description=_("Is full"))
    def is_full(self, obj):
        return obj.is_full

    @admin.display(boolean=True, description=_("Special price"))
    def special_price(self, obj):
        return obj.price.type == 1

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser or obj and not obj.event.ended

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser or obj and not obj.event.ended


@admin.register(SeminarRegistration)
class SeminarRegistrationAdmin(admin.ModelAdmin):
    list_display = ["teacher", "dancer", "academy", "seminar_date_time", "deposit_paid"]
    list_filter = ["seminar__event", "seminar__teacher"]
    list_display_links = ["dancer"]
    search_fields = ["academy__name", "dancer__first_name", "dancer__last_name"]
    search_help_text = _("Search by academy, dancer's first or last name.")
    show_facets = admin.ShowFacets.ALWAYS

    actions = ["export_excel"]

    @admin.action(description=_("Export to Excel"))
    def export_excel(self, request, queryset):
        excel_file = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = gettext("Seminar registrations list")

        headers = [
            gettext("Academy"),
            gettext("Dancer ID"),
            gettext("Dancer name"),
            gettext("Seminars"),
            gettext("Price per seminar"),
            gettext("Total"),
            gettext("Paid amount"),
            gettext("Balance"),
        ]

        headers_row = 1
        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=headers_row, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.font = Font(bold=True)

        dancers_qs = Dancer.objects.filter(
            seminar_registrations__in=queryset
        ).distinct()

        current_row = headers_row + 1
        for dancer in dancers_qs:
            (
                registrations,
                seminar_prices,
                total_amount,
                paid_amount,
                balance,
            ) = ([], [], 0, 0, 0)
            for seminar_registration in dancer.seminar_registrations.all():
                registrations.append(seminar_registration.seminar.__str__())
                seminar_prices.append(str(seminar_registration.total_price))
                total_amount += seminar_registration.total_price
                paid_amount -= seminar_registration.paid_amount
                balance += seminar_registration.balance

            row = [
                dancer.academy.__str__(),
                dancer.pk,
                dancer.__str__(),
                "\n".join(registrations),
                "\n".join(seminar_prices),
                total_amount,
                paid_amount,
                balance,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(vertical="center")

            worksheet.row_dimensions[current_row].height = len(registrations) * 15

            current_row += 1

        for column in worksheet.columns:
            for col in column:
                if hasattr(col, "column_letter"):
                    worksheet.column_dimensions[col.column_letter].width = 20

        workbook.save(excel_file)
        response = HttpResponse(
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response[
            "Content-Disposition"
        ] = "attachment; filename=Seminar registrations list.xlsx"
        return response

    fieldsets = (
        (
            None,
            {
                "fields": (
                    ("academy",),
                    ("seminar", "dancer"),
                    ("total_price", "balance"),
                    ("deposit_paid", "fully_paid"),
                )
            },
        ),
    )

    def get_readonly_fields(self, request, obj):
        readonly_fields = ["total_price", "deposit_paid", "fully_paid", "balance"]
        if not request.user.is_superuser and obj:
            readonly_fields.extend([field.name for field in self.model._meta.fields])
        return readonly_fields

    class SeminarPaymentInline(admin.TabularInline):
        model = SeminarPayment
        fk_name = "seminar_registration"
        extra = 0
        formset = SeminarPaymentFormSet

        def has_change_permission(self, request, obj=None):
            return (
                request.user.is_superuser or obj and not obj.seminar.registration_ended
            )

        def has_add_permission(self, request, obj=None):
            return (
                request.user.is_superuser
                or obj
                and not obj.seminar.registration_ended
                and obj.balance > 0
            )

        def has_delete_permission(self, request, obj=None):
            return (
                request.user.is_superuser or obj and not obj.seminar.registration_ended
            )

    inlines = [SeminarPaymentInline]

    @admin.display(description=_("Teacher"))
    def teacher(self, obj):
        return obj.seminar.teacher

    @admin.display(description=_("Date and time"))
    def seminar_date_time(self, obj):
        return f"{obj.seminar.date} {obj.seminar.time}"

    @admin.display(boolean=True, description=_("Fully paid"))
    def fully_paid(self, obj):
        return obj.fully_paid

    @admin.display(boolean=True, description=_("Deposit paid"))
    def deposit_paid(self, obj):
        return obj.deposit_paid

    @admin.display(description=_("Total price"))
    def total_price(self, obj):
        return f"$ {obj.total_price}"

    @admin.display(description=_("Balance"))
    def balance(self, obj):
        return f"$ {obj.balance}"

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser or obj and not obj.seminar.registration_ended


@admin.register(SeminarPayment)
class SeminarPaymentAdmin(admin.ModelAdmin):
    list_display = ["date", "seminar", "dancer", "amount", "payment_method"]
    list_filter = [
        "seminar_registration__seminar__event",
        "payment_method",
        "date",
    ]
    list_display_links = ["seminar", "amount"]
    search_fields = [
        "seminar_registration__academy__name",
        "date",
        "seminar_registration__dancer__first_name",
        "seminar_registration__dancer__last_name",
    ]
    search_help_text = _("Search by academy, date, dancer's first or last name.")
    show_facets = admin.ShowFacets.ALWAYS

    actions = ["export_excel"]

    @admin.display(description=_("Seminar"))
    def seminar(self, obj):
        return obj.seminar_registration.seminar

    @admin.display(description=_("Dancer"))
    def dancer(self, obj):
        return obj.seminar_registration.dancer

    @admin.action(description=_("Export to Excel"))
    def export_excel(self, request, queryset):
        excel_file = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = gettext("Seminar payments list")

        headers = [
            gettext("Date"),
            gettext("Teacher"),
            gettext("Dancer"),
            gettext("Academy"),
            gettext("Amount"),
            gettext("Payment method"),
        ]

        total_amount = 0
        for payment in queryset:
            total_amount += payment.amount

        title_cell = worksheet.cell(row=1, column=1)
        count = queryset.count()
        message = ngettext(
            "%(count)d selected payment - Total $ %(amount)s",
            "%(count)d selected payments - Total $ %(amount)s",
            count,
        ) % {
            "count": count,
            "amount": total_amount,
        }
        title_cell.value = message
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.merge_cells(
            start_column=1, end_column=len(headers), start_row=1, end_row=1
        )

        headers_row = 2
        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=headers_row, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.font = Font(bold=True)

        current_row = headers_row + 1
        for payment in queryset:
            row = [
                payment.date,
                payment.seminar_registration.seminar.__str__(),
                payment.seminar_registration.student.__str__(),
                payment.seminar_registration.academy.__str__(),
                payment.amount,
                payment.get_payment_method_display(),
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(vertical="center")

            current_row += 1

        for column in worksheet.columns:
            for col in column:
                if hasattr(col, "column_letter"):
                    worksheet.column_dimensions[col.column_letter].width = 20

        workbook.save(excel_file)
        response = HttpResponse(
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response[
            "Content-Disposition"
        ] = "attachment; filename=Seminar payments list.xlsx"
        return response

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "seminar_registration",
                    "amount",
                    "payment_method",
                    "date",
                )
            },
        ),
    )

    form = SeminarPaymentAdminForm

    def get_readonly_fields(self, request, obj):
        return ["seminar_registration"] if obj else []

    def has_change_permission(self, request, obj=None):
        return (
            request.user.is_superuser
            or obj
            and not obj.seminar_registration.seminar.registration_ended
        )

    def has_delete_permission(self, request, obj=None):
        return (
            request.user.is_superuser
            or obj
            and not obj.seminar_registration.seminar.registration_ended
        )
